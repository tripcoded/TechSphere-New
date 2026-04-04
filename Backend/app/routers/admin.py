import re
from datetime import datetime, timezone
from io import BytesIO
from typing import Literal
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException, Query, Response, status
from sqlalchemy import select
from sqlalchemy.orm import Session, selectinload

from app.database import get_db
from app.dependencies import require_admin_api_key
from app.models import Attendance, AttendanceStatus, Event, Team, TeamMember, User
from app.profile_utils import has_completed_academic_profile

router = APIRouter(prefix="/admin", tags=["Admin"])


@router.get("/validate", dependencies=[Depends(require_admin_api_key)])
def validate_admin_key():
    return {"ok": True}


def _export_filename(event: Event, format_name: Literal["xlsx", "pdf"]) -> str:
    safe_title = re.sub(r'[\\/:*?"<>|]+', " ", (event.title or "")).strip()
    safe_title = re.sub(r"\s+", " ", safe_title).strip(". ")
    return f"{safe_title or f'Event {event.id}'} participants.{format_name}"


def _format_datetime(value: datetime | None) -> str:
    if value is None:
        return "-"
    if value.tzinfo is not None:
        return value.astimezone(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    return value.strftime("%Y-%m-%d %H:%M")


def _serialize_event_export_data(db: Session, event_id: int) -> tuple[Event, list[dict[str, object]], int]:
    event = db.get(Event, event_id)
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Event not found")

    teams = db.scalars(
        select(Team)
        .where(Team.event_id == event_id)
        .options(
            selectinload(Team.members).selectinload(TeamMember.user).selectinload(User.profile),
            selectinload(Team.leader).selectinload(User.profile),
        )
        .order_by(Team.created_at.asc())
    ).all()

    attendance_rows = db.scalars(select(Attendance).where(Attendance.event_id == event_id)).all()
    attendance_map = {row.user_id: row.status.value for row in attendance_rows}

    team_sections: list[dict[str, object]] = []
    total_participants = 0
    for team in teams:
        ordered_memberships = sorted(team.members, key=lambda membership: membership.joined_at)
        members: list[dict[str, object]] = []
        for index, membership in enumerate(ordered_memberships, start=1):
            user = membership.user
            profile = user.profile
            digital_presence_items = []
            if profile and profile.github_url:
                digital_presence_items.append(f"GitHub: {profile.github_url}")
            if profile and profile.linkedin_url:
                digital_presence_items.append(f"LinkedIn: {profile.linkedin_url}")
            if profile and profile.portfolio_url:
                digital_presence_items.append(f"Portfolio: {profile.portfolio_url}")

            members.append(
                {
                    "participant_order": index,
                    "team_role": "Leader" if membership.user_id == team.leader_id else "Member",
                    "participant_name": user.full_name or "",
                    "participant_email": user.email,
                    "roll_no": profile.roll_no if profile else "",
                    "branch": profile.branch if profile else "",
                    "year": profile.year if profile and profile.year is not None else "",
                    "academic_profile_completed": "Yes" if has_completed_academic_profile(profile) else "No",
                    "github_url": profile.github_url if profile else "",
                    "linkedin_url": profile.linkedin_url if profile else "",
                    "portfolio_url": profile.portfolio_url if profile else "",
                    "headline": profile.headline if profile else "",
                    "skills": profile.skills if profile else "",
                    "bio": profile.bio if profile else "",
                    "digital_presence": "\n".join(digital_presence_items) if digital_presence_items else "Not added",
                    "attendance": attendance_map.get(user.id, AttendanceStatus.absent.value),
                    "joined_at": _format_datetime(membership.joined_at),
                }
            )

        total_participants += len(members)
        leader_name = team.leader.full_name if team.leader and team.leader.full_name else team.leader.email if team.leader else "-"
        team_sections.append(
            {
                "team_name": team.name,
                "team_created_at": _format_datetime(team.created_at),
                "leader_name": leader_name,
                "member_count": len(members),
                "members": members,
            }
        )

    return event, team_sections, total_participants


def _build_excel_export(event: Event, team_sections: list[dict[str, object]], total_participants: int) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Excel export dependencies are not installed on the server.",
        ) from exc

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Participants"

    title_fill = PatternFill("solid", fgColor="0F172A")
    section_fill = PatternFill("solid", fgColor="1D4ED8")
    meta_fill = PatternFill("solid", fgColor="EFF6FF")
    header_fill = PatternFill("solid", fgColor="1E40AF")
    header_font = Font(color="FFFFFF", bold=True)
    title_font = Font(color="FFFFFF", bold=True, size=18)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1"),
    )

    sheet.merge_cells("A1:J1")
    sheet["A1"] = f"{event.title} Participant Export"
    sheet["A1"].fill = title_fill
    sheet["A1"].font = title_font
    sheet["A1"].alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 28

    summary_rows = [
        ("Event Name", event.title),
        ("Event ID", event.id),
        ("Description", event.description or "No description provided."),
        ("Location", event.location or "TBA"),
        ("Starts At", _format_datetime(event.starts_at)),
        ("Ends At", _format_datetime(event.ends_at)),
        ("Registered Teams", len(team_sections)),
        ("Registered Participants", total_participants),
    ]

    current_row = 3
    for field, value in summary_rows:
        sheet.cell(row=current_row, column=1, value=field).fill = header_fill
        sheet.cell(row=current_row, column=1).font = header_font
        sheet.cell(row=current_row, column=2, value=value).fill = meta_fill
        sheet.cell(row=current_row, column=1).alignment = wrap_alignment
        sheet.cell(row=current_row, column=2).alignment = wrap_alignment
        sheet.cell(row=current_row, column=1).border = thin_border
        sheet.cell(row=current_row, column=2).border = thin_border
        current_row += 1

    current_row += 2
    member_headers = [
        "Member Name",
        "Role",
        "Email",
        "Roll No",
        "Branch",
        "Year",
        "Digital Presence",
        "Academic Profile",
        "Attendance",
        "Joined At",
    ]

    if not team_sections:
        sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
        empty_cell = sheet.cell(row=current_row, column=1, value="No teams or participants are registered for this event yet.")
        empty_cell.fill = meta_fill
        empty_cell.alignment = wrap_alignment
    else:
        for team in team_sections:
            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            team_cell = sheet.cell(
                row=current_row,
                column=1,
                value=f"Team Name: {team['team_name']}",
            )
            team_cell.fill = section_fill
            team_cell.font = header_font
            team_cell.alignment = wrap_alignment
            current_row += 1

            sheet.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=10)
            meta_cell = sheet.cell(
                row=current_row,
                column=1,
                value=f"Leader: {team['leader_name']} | Members: {team['member_count']} | Created At: {team['team_created_at']}",
            )
            meta_cell.fill = meta_fill
            meta_cell.alignment = wrap_alignment
            current_row += 1

            for index, header in enumerate(member_headers, start=1):
                header_cell = sheet.cell(row=current_row, column=index, value=header)
                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = wrap_alignment
                header_cell.border = thin_border
            current_row += 1

            for member in team["members"]:
                values = [
                    member["participant_name"] or member["participant_email"],
                    member["team_role"],
                    member["participant_email"],
                    member["roll_no"] or "-",
                    member["branch"] or "-",
                    member["year"] or "-",
                    member["digital_presence"],
                    member["academic_profile_completed"],
                    member["attendance"],
                    member["joined_at"],
                ]
                for index, value in enumerate(values, start=1):
                    data_cell = sheet.cell(row=current_row, column=index, value=value)
                    data_cell.alignment = wrap_alignment
                    data_cell.border = thin_border
                current_row += 1

            current_row += 1

    column_widths = {
        "A": 24,
        "B": 12,
        "C": 30,
        "D": 18,
        "E": 28,
        "F": 10,
        "G": 42,
        "H": 18,
        "I": 14,
        "J": 24,
    }
    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _build_pdf_export(event: Event, team_sections: list[dict[str, object]], total_participants: int) -> bytes:
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, TableStyle
    except ImportError as exc:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="PDF export dependencies are not installed on the server.",
        ) from exc

    buffer = BytesIO()
    document = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=24,
        rightMargin=24,
        topMargin=24,
        bottomMargin=24,
    )

    styles = getSampleStyleSheet()
    small = ParagraphStyle(
        "SmallWrap",
        parent=styles["BodyText"],
        fontSize=8,
        leading=10,
        spaceAfter=0,
    )

    story = [
        Paragraph(f"{event.title} Participant Export", styles["Title"]),
        Spacer(1, 8),
        Paragraph(f"<b>Event Name:</b> {event.title}", styles["BodyText"]),
        Paragraph(f"<b>Event ID:</b> {event.id}", styles["BodyText"]),
        Paragraph(f"<b>Location:</b> {event.location or 'TBA'}", styles["BodyText"]),
        Paragraph(f"<b>Starts At:</b> {_format_datetime(event.starts_at)}", styles["BodyText"]),
        Paragraph(f"<b>Ends At:</b> {_format_datetime(event.ends_at)}", styles["BodyText"]),
        Paragraph(f"<b>Description:</b> {event.description or 'No description provided.'}", styles["BodyText"]),
        Paragraph(f"<b>Registered Teams:</b> {len(team_sections)}", styles["BodyText"]),
        Paragraph(f"<b>Registered Participants:</b> {total_participants}", styles["BodyText"]),
        Spacer(1, 14),
    ]

    if team_sections:
        for team in team_sections:
            story.append(Paragraph(f"Team: {team['team_name']}", styles["Heading2"]))
            story.append(
                Paragraph(
                    f"<b>Leader:</b> {team['leader_name']} &nbsp;&nbsp;&nbsp; <b>Members:</b> {team['member_count']} &nbsp;&nbsp;&nbsp; <b>Created At:</b> {team['team_created_at']}",
                    styles["BodyText"],
                )
            )
            story.append(Spacer(1, 8))

            table_data = [
                [
                    Paragraph("<b>Member</b>", small),
                    Paragraph("<b>Role</b>", small),
                    Paragraph("<b>Email</b>", small),
                    Paragraph("<b>Roll No</b>", small),
                    Paragraph("<b>Branch</b>", small),
                    Paragraph("<b>Year</b>", small),
                    Paragraph("<b>Digital Presence</b>", small),
                    Paragraph("<b>Attendance</b>", small),
                ]
            ]

            for member in team["members"]:
                digital_presence = str(member["digital_presence"]).replace("\n", "<br/>")
                table_data.append(
                    [
                        Paragraph(str(member["participant_name"] or member["participant_email"] or "-"), small),
                        Paragraph(str(member["team_role"] or "-"), small),
                        Paragraph(str(member["participant_email"] or "-"), small),
                        Paragraph(str(member["roll_no"] or "-"), small),
                        Paragraph(str(member["branch"] or "-"), small),
                        Paragraph(str(member["year"] or "-"), small),
                        Paragraph(digital_presence or "-", small),
                        Paragraph(str(member["attendance"] or "-"), small),
                    ]
                )

            table = LongTable(
                table_data,
                repeatRows=1,
                colWidths=[96, 52, 146, 62, 96, 34, 200, 62],
            )
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1D4ED8")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F8FAFC")),
                        ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CBD5E1")),
                        ("VALIGN", (0, 0), (-1, -1), "TOP"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                        ("TOPPADDING", (0, 0), (-1, -1), 5),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ]
                )
            )
            story.append(table)
            story.append(Spacer(1, 14))
    else:
        story.append(Paragraph("No participants are registered for this event yet.", styles["BodyText"]))

    document.build(story)
    buffer.seek(0)
    return buffer.getvalue()


@router.get("/events/{event_id}/participants/export", dependencies=[Depends(require_admin_api_key)])
def export_event_participants(
    event_id: int,
    format: Literal["xlsx", "pdf"] = Query(default="xlsx"),
    db: Session = Depends(get_db),
):
    event, team_sections, total_participants = _serialize_event_export_data(db, event_id)
    filename = _export_filename(event, format)
    content_disposition = f"attachment; filename=\"{filename}\"; filename*=UTF-8''{quote(filename)}"

    if format == "pdf":
        content = _build_pdf_export(event, team_sections, total_participants)
        media_type = "application/pdf"
    else:
        content = _build_excel_export(event, team_sections, total_participants)
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    return Response(
        content=content,
        media_type=media_type,
        headers={"Content-Disposition": content_disposition},
    )
